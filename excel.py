from openpyxl import Workbook


class Excel:
    def __init__(self,
                 file_name: str,
                 sheet_names: list) -> None:

        self.wb = Workbook()
        self.file_name = file_name

        for sheet_name in sheet_names:
            self.wb.create_sheet(sheet_name)

    def save_data(self,
                  sheet_name,
                  column_name: str,
                  column_sequence: int,
                  user_names: list,
                  user_phone_numbers: list):

        ws = self.wb[sheet_name]

        ws.merge_cells(start_row=1, start_column=column_sequence * 2 + 1, end_row=1, end_column=column_sequence * 2 + 2)

        ws.cell(row=1, column=column_sequence * 2 + 1).value = column_name

        row = 2

        for user_name in user_names:
            ws.cell(row=row, column=column_sequence * 2 + 1).value = user_name
            row += 1

        row = 2

        for user_phone_number in user_phone_numbers:
            ws.cell(row=row, column=column_sequence * 2 + 2).value = user_phone_number
            row += 1

    def save(self):
        self.wb.save(f'{self.file_name}.xlsx')
