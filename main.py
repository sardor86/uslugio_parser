from config import BaseDriver, path, \
    URL, BROWSER_SETTINGS, LOG_IN_URL

from requests import get
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import Workbook

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException


class Excel:
    def __init__(self,
                 file_name: str,
                 sheet_names: list) -> None:

        self.wb = Workbook()
        self.file_name = file_name

        for sheet_name in sheet_names:
            self.wb.create_sheet(sheet_name)

    def save_data(self,
                  sheet_name,
                  column_name: str,
                  column_sequence: int,
                  user_names: list,
                  user_phone_numbers: list):

        ws = self.wb[sheet_name]

        ws.merge_cells(start_row=1, start_column=column_sequence * 2 + 1, end_row=1, end_column=column_sequence * 2 + 2)

        ws.cell(row=1, column=column_sequence * 2 + 1).value = column_name

        row = 2

        for user_name in user_names:
            ws.cell(row=row, column=column_sequence * 2 + 1).value = user_name
            row += 1

        row = 2

        for user_phone_number in user_phone_numbers:
            ws.cell(row=row, column=column_sequence * 2 + 2).value = user_phone_number
            row += 1

    def save(self):
        self.wb.save(f'{self.file_name}.xlsx')


class Driver(BaseDriver):
    def scroll_down(self, down_counter: int) -> None:
        for i in range(down_counter):
            self.find_element(By.TAG_NAME, "html").send_keys(Keys.DOWN)

    def get_phone_and_name_service(self):
        self.get(self.url)
        self.scroll_down(1000)

        item_data = self.find_elements(By.CLASS_NAME, 'item_data')
        item_bottom = self.find_elements(By.CLASS_NAME, 'item-bottom')

        for data in item_data:
            self.user_name.append(data.find_element(By.CLASS_NAME, 'phoneline').find_element(By.TAG_NAME, 'span').text)

        for bottom in item_bottom:
            bottom.find_element(By.CLASS_NAME, 'showphone1').click()

            while True:
                try:
                    self.phone_number.append(bottom.find_element(By.TAG_NAME, 'a').text)
                    sleep(1)
                    break
                except NoSuchElementException:
                    sleep(1)

    def log_in(self, email: str, password: str) -> bool:
        self.get(LOG_IN_URL)
        sleep(3)

        self.find_element(By.XPATH, '/html/body/div[2]/div/div/div[1]/div/form/div[1]/input').send_keys(email)
        self.find_element(By.XPATH, '/html/body/div[2]/div/div/div[1]/div/form/div[2]/input').send_keys(password)

        self.find_element(By.XPATH, '/html/body/div[2]/div/div/div[1]/div/form/button').click()

        try:
            self.find_element(By.XPATH, '/html/body/div[2]/div/div/div[1]/div/form/div[3]/ul/li')
            return False
        except NoSuchElementException:
            sleep(2)
            return True


def get_all_countries() -> list:
    content = get(URL,
                  headers={'User-Agent': BROWSER_SETTINGS}, )
    page = BeautifulSoup(content.content, 'html.parser')

    col_countries = page.find_all('div', class_='col-sm-4')
    all_countries = []

    for colm in col_countries:
        for countries in colm.find_all('a'):
            all_countries.append((countries.get_text(), URL + countries.get('href')))

    return all_countries


def get_all_services(url: str) -> list:
    content = get(url,
                  headers={'User-Agent': BROWSER_SETTINGS}, )
    page = BeautifulSoup(content.content, 'html.parser')

    services = page.find('div', class_='nav_cat_list')

    return [(service.text, URL + service.get('href')) for service in services.find_all('a')]


def main():
    index = 0

    for country in get_all_countries():
        print(f'{index}) {country[0]}')
        index += 1

    print(f'{index}) все')

    country_number = input('Выберите город: ')

    if not country_number.isdigit():
        print('Вы неправильно написали номер города')
        sleep(5)
        return

    country_number = int(country_number)

    if country_number > index:
        print('такого города нет')
        sleep(5)
        return

    if country_number == index:
        for country in get_all_countries():
            for services in get_all_services(country[1]):
                print(services)

                return

    country = get_all_countries()[country_number]

    index = 0
    for service in get_all_services(country[1]):
        print(f'{index}) {service[0]}')
        index += 1

    print(f'{index}) все')

    service_number = input('Выберите сервис: ')

    if not service_number.isdigit():
        print('Вы неправильно написали номер города')
        sleep(5)
        return

    service_number = int(service_number)

    if service_number > index:
        print('такого сервиса нет')
        sleep(5)
        return

    if service_number == index:
        for services in get_all_services(country[1]):
            print(services)
    else:
        print(get_all_services(country[1])[service_number])


if __name__ == '__main__':
    main()
